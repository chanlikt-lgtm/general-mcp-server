import csv
import io
import json
import os
from mcp.server.fastmcp import FastMCP

_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(_BASE, "data")


def _safe(filename: str) -> str | None:
    target = os.path.normpath(os.path.join(DATA_DIR, filename))
    return target if target.startswith(DATA_DIR) else None


def register_pdf_office_tools(mcp: FastMCP) -> None:

    @mcp.tool()
    def pdf_to_text(filename: str, pages: str = "") -> str:
        """
        Extract text from a PDF file in the data/ directory.
        filename: relative path, e.g. 'document.pdf'.
        pages: optional page range, e.g. '1-3' or '2' (default: all pages).
        Returns extracted text with page separators.
        """
        from pypdf import PdfReader
        path = _safe(filename)
        if not path: return "Error: path traversal not allowed."
        if not os.path.exists(path): return f"File not found: {filename}"
        try:
            reader = PdfReader(path)
            total = len(reader.pages)
            if pages:
                parts = pages.split("-")
                start = int(parts[0]) - 1
                end   = int(parts[1]) if len(parts) > 1 else int(parts[0])
                page_range = range(max(0, start), min(total, end))
            else:
                page_range = range(total)
            result = []
            for i in page_range:
                text = reader.pages[i].extract_text() or ""
                result.append(f"--- Page {i+1} ---\n{text.strip()}")
            return "\n\n".join(result) or "No text extracted."
        except Exception as e:
            return f"Error: {e}"

    @mcp.tool()
    def pdf_info(filename: str) -> str:
        """
        Get metadata about a PDF file: page count, title, author, creation date.
        filename: relative path in data/, e.g. 'report.pdf'.
        """
        from pypdf import PdfReader
        path = _safe(filename)
        if not path: return "Error: path traversal not allowed."
        if not os.path.exists(path): return f"File not found: {filename}"
        try:
            reader = PdfReader(path)
            meta = reader.metadata or {}
            return (
                f"Pages   : {len(reader.pages)}\n"
                f"Title   : {meta.get('/Title', 'N/A')}\n"
                f"Author  : {meta.get('/Author', 'N/A')}\n"
                f"Created : {meta.get('/CreationDate', 'N/A')}\n"
                f"Subject : {meta.get('/Subject', 'N/A')}"
            )
        except Exception as e:
            return f"Error: {e}"

    @mcp.tool()
    def merge_pdfs(filenames: str, output: str) -> str:
        """
        Merge multiple PDF files into one.
        filenames: comma-separated list of PDF paths in data/, e.g. 'a.pdf,b.pdf,c.pdf'.
        output: output filename in data/, e.g. 'merged.pdf'.
        Returns confirmation with page count.
        """
        from pypdf import PdfWriter, PdfReader
        out_path = _safe(output)
        if not out_path: return "Error: invalid output path."
        writer = PdfWriter()
        total_pages = 0
        for fname in [f.strip() for f in filenames.split(",")]:
            path = _safe(fname)
            if not path: return f"Error: invalid path '{fname}'"
            if not os.path.exists(path): return f"File not found: {fname}"
            try:
                reader = PdfReader(path)
                for page in reader.pages:
                    writer.add_page(page)
                total_pages += len(reader.pages)
            except Exception as e:
                return f"Error reading '{fname}': {e}"
        try:
            os.makedirs(os.path.dirname(out_path), exist_ok=True)
            with open(out_path, "wb") as f:
                writer.write(f)
            return f"Merged {len(filenames.split(','))} PDFs → {output} ({total_pages} pages)"
        except Exception as e:
            return f"Error writing output: {e}"

    @mcp.tool()
    def split_pdf(filename: str, pages_per_file: int = 1) -> str:
        """
        Split a PDF into multiple files, each containing a fixed number of pages.
        filename: source PDF in data/, e.g. 'document.pdf'.
        pages_per_file: number of pages per output file (default 1 = one file per page).
        Output files are saved as 'document_part1.pdf', 'document_part2.pdf', etc.
        Returns list of created files.
        """
        from pypdf import PdfReader, PdfWriter
        path = _safe(filename)
        if not path: return "Error: path traversal not allowed."
        if not os.path.exists(path): return f"File not found: {filename}"
        pages_per_file = max(1, int(pages_per_file))
        try:
            reader = PdfReader(path)
            total  = len(reader.pages)
            base   = os.path.splitext(filename)[0]
            created = []
            part = 1
            for start in range(0, total, pages_per_file):
                writer = PdfWriter()
                for i in range(start, min(start + pages_per_file, total)):
                    writer.add_page(reader.pages[i])
                out_name = f"{base}_part{part}.pdf"
                out_path = _safe(out_name)
                if not out_path: continue
                with open(out_path, "wb") as f:
                    writer.write(f)
                created.append(out_name)
                part += 1
            return f"Split into {len(created)} files:\n" + "\n".join(created)
        except Exception as e:
            return f"Error: {e}"

    @mcp.tool()
    def csv_to_excel(csv_filename: str, excel_filename: str = "") -> str:
        """
        Convert a CSV file to an Excel (.xlsx) file.
        csv_filename: source CSV in data/, e.g. 'data.csv'.
        excel_filename: output Excel filename (default: same name with .xlsx extension).
        Returns confirmation with row/column count.
        """
        import openpyxl
        path = _safe(csv_filename)
        if not path: return "Error: path traversal not allowed."
        if not os.path.exists(path): return f"File not found: {csv_filename}"
        if not excel_filename:
            excel_filename = os.path.splitext(csv_filename)[0] + ".xlsx"
        out_path = _safe(excel_filename)
        if not out_path: return "Error: invalid output path."
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            with open(path, encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                rows = list(reader)
                for row in rows:
                    ws.append(row)
            wb.save(out_path)
            cols = len(rows[0]) if rows else 0
            return f"Converted {csv_filename} → {excel_filename} ({len(rows)} rows × {cols} cols)"
        except Exception as e:
            return f"Error: {e}"

    @mcp.tool()
    def excel_to_csv(excel_filename: str, csv_filename: str = "", sheet: str = "") -> str:
        """
        Convert an Excel (.xlsx) file to CSV.
        excel_filename: source Excel file in data/, e.g. 'report.xlsx'.
        csv_filename: output CSV filename (default: same name with .csv extension).
        sheet: sheet name to export (default: first sheet).
        Returns confirmation with row/column count.
        """
        import openpyxl
        path = _safe(excel_filename)
        if not path: return "Error: path traversal not allowed."
        if not os.path.exists(path): return f"File not found: {excel_filename}"
        if not csv_filename:
            csv_filename = os.path.splitext(excel_filename)[0] + ".csv"
        out_path = _safe(csv_filename)
        if not out_path: return "Error: invalid output path."
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            rows = list(ws.iter_rows(values_only=True))
            with open(out_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerows([[str(c) if c is not None else "" for c in row] for row in rows])
            cols = len(rows[0]) if rows else 0
            return f"Converted {excel_filename} (sheet: {ws.title}) → {csv_filename} ({len(rows)} rows × {cols} cols)"
        except Exception as e:
            return f"Error: {e}"
